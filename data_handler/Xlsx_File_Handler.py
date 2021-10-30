import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class Xlsx_File_Handler:
    xlsx_file = None
    xlsx_sheet = None
    absolute_file_path = None
    row_index = 1
    unmatched_data_fill_style = None

    def __init__(self, file_name, headers_list):
        self.absolute_file_path = self.get_complete_xlsx_file_path(file_name)
        self.xlsx_file = Workbook()
        self.xlsx_sheet = self.xlsx_file.active
        self.xlsx_sheet.title = 'Comparison_Result'
        self.number_of_cols = len(headers_list)
        self.write_row_in_xlsx_file(headers_list, Font(bold=True))
        self.unmatched_data_fill_style = PatternFill(start_color='FFFFA500',
                                                     end_color='FFFFA500',
                                                     fill_type='solid')
        self.blank_row_fill_style = PatternFill(start_color='FFC39BD3',
                                                end_color='FFC39BD3',
                                                fill_type='solid')
        self.failure_font = Font(color='FF0000')

    def get_complete_xlsx_file_path(self, file_name):
        project_root = Path(os.path.abspath(os.path.dirname(__file__))).parent
        file_dir = 'report/xlsx_output'
        file_dir = os.path.join(project_root, file_dir)
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)
        file_path = os.path.join(file_dir, file_name)
        return file_path

    def write_row_in_xlsx_file(self, list_of_col_values_for_a_row: list, font=None):
        for col_num, val in enumerate(list_of_col_values_for_a_row, start=1):
            self.xlsx_sheet.cell(row=self.row_index, column=col_num).value = val
        if font:
            for col_num, val in enumerate(list_of_col_values_for_a_row, start=1):
                self.xlsx_sheet.cell(row=self.row_index, column=col_num).font = font
        self.row_index += 1

    def write_blank_colored_row_in_xlsx_file(self):
        for col_number in range(1, self.number_of_cols + 1):
            self.xlsx_sheet.cell(row=self.row_index, column=col_number).fill = self.blank_row_fill_style
        self.row_index += 1

    def write_partition1_and_partition2_row_in_xlsx_file(self, partition_1_row: tuple,
                                                         partition_2_row: tuple,
                                                         unmatched_col_indexes: list = None):
        self.write_row_in_xlsx_file(list(partition_1_row))
        if unmatched_col_indexes:
            for index in unmatched_col_indexes:
                self.xlsx_sheet.cell(row=self.row_index,
                                     column=index+1).fill = self.unmatched_data_fill_style
                self.xlsx_sheet.cell(row=self.row_index,
                                     column=index + 1).font = self.failure_font
        self.write_row_in_xlsx_file(list(partition_2_row))
        self.write_blank_colored_row_in_xlsx_file()

    def save_xlsx_file(self):
        # check if workbook is opened by some other process. Close the workbook in that case

        print('Saving xlsx file ... ')
        self.xlsx_file.save(filename=self.absolute_file_path)
        self.xlsx_file.close()
        print('XLSX file saved successfully!!')


if __name__ == '__main__':
    fh = Xlsx_File_Handler('testfile.xlsx', ['col1', 'col2', 'col3'])
    fh.save_xlsx_file()
